import asyncio
import csv
import datetime
import threading
from concurrent.futures import ThreadPoolExecutor

import openpyxl
from PyQt5 import QtCore
from uitls.logger import logger
from uitls.DBconnectServer import popup_manager


class ExportThread(threading.Thread):
    def __init__(self, data, headers, file_path, chunk_size=50000, num_threads=10):
        threading.Thread.__init__(self)
        self.data = data
        self.headers = headers
        self.file_path = file_path
        self.chunk_size = chunk_size
        self.num_threads = num_threads
        self.export_finished = False  # 导出完成标志

    def run(self):
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        loop.run_until_complete(self.write_to_excel())

    async def write_to_excel(self):

        workbook = openpyxl.Workbook()

        worksheet = workbook.active
        StartTime = datetime.datetime.now()

        for col_idx, header in enumerate(self.headers, start=1):
            worksheet.cell(row=1, column=col_idx, value=str(header))

        with ThreadPoolExecutor(max_workers=self.num_threads) as executor:
            futures = []
            for chunk_start in range(0, len(self.data), self.chunk_size):
                chunk_end = min(chunk_start + self.chunk_size, len(self.data))
                chunk_data = self.data[chunk_start:chunk_end]
                future = executor.submit(self.write_chunk_to_excel, worksheet, chunk_data, chunk_start + 2)
                futures.append(future)

            await asyncio.gather(*[asyncio.wrap_future(f) for f in futures])

            workbook.save(self.file_path)
            workbook.close()
            EndTime = datetime.datetime.now()
            DiffTime = EndTime - StartTime

            logger.info(f"{datetime.datetime.now()} ------------导出完成-----------------")
            logger.info(f"导出完成,耗时：{round(DiffTime.total_seconds(), 4)}秒")
            print(f"{datetime.datetime.now()} ------------导出完成-----------------")
            print(f"导出完成,耗时：{round(DiffTime.total_seconds(), 4)}秒")
            popup_manager.message_info.emit(f"导出完成,耗时：{round(DiffTime.total_seconds(), 4)}秒")
            self.export_finished = True
    def write_chunk_to_excel(self, worksheet, chunk_data, start_row):
        for row_idx, row_data in enumerate(chunk_data, start=start_row):
            for col_idx, cell_data in enumerate(row_data, start=1):
                worksheet.cell(row=row_idx, column=col_idx, value=str(cell_data))

class ExportThreadCsv(threading.Thread):
    def __init__(self, data, headers, file_path, chunk_size=50000, num_threads=10):
        threading.Thread.__init__(self)
        self.data = data
        self.headers = headers
        self.file_path = file_path
        self.chunk_size = chunk_size
        self.num_threads = num_threads
        self.export_finished = False  # 导出完成标志

    def run(self):
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        loop.run_until_complete(self.write_to_csv())

    async def write_to_csv(self):
        StartTime = datetime.datetime.now()
        logger.info(f"{datetime.datetime.now()} ------------开始导出 CSV -----------------")
        print(f"{datetime.datetime.now()} ------------开始导出 CSV -----------------")

        # 打开文件进行写入
        with open(self.file_path, mode='w', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            # 写入表头
            writer.writerow(self.headers)
            # 准备线程池
            with ThreadPoolExecutor(max_workers=self.num_threads) as executor:
                futures = []
                for chunk_start in range(0, len(self.data), self.chunk_size):
                    chunk_end = min(chunk_start + self.chunk_size, len(self.data))
                    chunk_data = self.data[chunk_start:chunk_end]
                    future = executor.submit(self.write_chunk_to_csv, writer, chunk_data)
                    futures.append(future)

                # 等待所有线程完成
                await asyncio.gather(*[asyncio.wrap_future(f) for f in futures])

        EndTime = datetime.datetime.now()
        DiffTime = EndTime - StartTime
        logger.info(f"{datetime.datetime.now()} ------------导出完成-----------------")
        logger.info(f"导出完成, 耗时：{round(DiffTime.total_seconds(), 4)}秒")
        print(f"{datetime.datetime.now()} ------------导出完成-----------------")
        print(f"导出完成, 耗时：{round(DiffTime.total_seconds(), 4)}秒")

        # 通知导出完成
        popup_manager.message_info.emit(f"导出完成, 耗时：{round(DiffTime.total_seconds(), 4)}秒")
        self.export_finished = True

    def write_chunk_to_csv(self, writer, chunk_data):
        # 将数据写入 CSV 文件
        for row_data in chunk_data:
            writer.writerow(row_data)
class LargeTableModel(QtCore.QAbstractTableModel):
    def __init__(self, data, headers, parent=None):
        super(LargeTableModel, self).__init__(parent)
        self._data = data if isinstance(data, list) else list(data)
        self._headers = headers
        self._sort_order = QtCore.Qt.AscendingOrder
        self._sort_column = -1

    def rowCount(self, parent=QtCore.QModelIndex()):
        return len(self._data)

    def columnCount(self, parent=QtCore.QModelIndex()):
        return len(self._headers)

    def data(self, index, role=QtCore.Qt.DisplayRole):
        if not index.isValid():
            return None

        if role == QtCore.Qt.DisplayRole:
            return str(self._data[index.row()][index.column()])
        return None

    def headerData(self, section, orientation, role=QtCore.Qt.DisplayRole):
        if role != QtCore.Qt.DisplayRole:
            return None

        if orientation == QtCore.Qt.Horizontal:
            return self._headers[section]
        elif orientation == QtCore.Qt.Vertical:
            return str(section + 1)
        return None

    def sort(self, column, order):
        self.layoutAboutToBeChanged.emit()
        self._data.sort(key=lambda row: row[column], reverse=(order == QtCore.Qt.DescendingOrder))
        self.layoutChanged.emit()

    def appendRow(self, row_data):
        """向模型中添加一行数据"""
        if isinstance(row_data, tuple):
            row_data = list(row_data)
        elif not isinstance(row_data, list):
            raise ValueError("row_data must be a list or tuple")

        self.beginInsertRows(QtCore.QModelIndex(), self.rowCount(), self.rowCount())
        self._data.append(row_data)
        self.endInsertRows()
