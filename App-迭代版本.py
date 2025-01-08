import asyncio
import base64
import csv
from concurrent.futures import ThreadPoolExecutor
import encodings.idna
import configparser
import datetime
import logging
import os
import queue
import re
import threading
from logging.handlers import TimedRotatingFileHandler
import openpyxl
import pandas as pd
# from sshtunnel import SSHTunnelForwarder
import pymysql
import sqlparse
from PyQt5 import QtCore, QtGui
from PyQt5 import QtWidgets
from PyQt5.QtCore import pyqtSignal, QThread, QEvent, QTimer, QTime
from PyQt5.QtCore import Qt, QItemSelection, QItemSelectionModel, QSortFilterProxyModel
from PyQt5.QtGui import QIcon, QStandardItemModel, QStandardItem
from PyQt5.QtWidgets import QFileDialog, QMenu, QAction, QLabel, QSplitter, QPushButton, \
    QComboBox, QAbstractItemView, QToolTip, QDialog, QFormLayout, QLineEdit, QCheckBox, QHBoxLayout, QVBoxLayout, \
    QListWidget, QTableWidget, QTableWidgetItem, QWidget, QListWidgetItem, QProgressBar
from PopupManager import *
import time

# 创建一个日志记录器
logger = logging.getLogger()
logger.setLevel(logging.INFO)
popup_manager = PopupManager()
if not os.path.exists('server'):
    os.makedirs('server')
if not os.path.exists('logs'):
    os.makedirs('logs')
# 创建一个处理程序，用于写入每天的日志到不同的文件
filename = datetime.datetime.now().strftime("%Y-%m-%d") + '.log'
log_path = os.path.join('logs', filename)
handler = TimedRotatingFileHandler(log_path, when="midnight", interval=1, backupCount=7)
handler.setLevel(logging.DEBUG)
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
handler.setFormatter(formatter)
# 将处理程序添加到日志记录器
logger.addHandler(handler)

global Process_df
Process_df = pd.DataFrame(columns=['Process_id', 'Status', '服务器组'])


def encode_password(password):
    """加密"""
    password_bytes = password.encode('utf-8')
    suffix_bytes = "aqa&".encode('utf-8')
    combined_bytes = password_bytes + suffix_bytes
    encoded_bytes = base64.b64encode(combined_bytes)
    return encoded_bytes.decode('utf-8')


def decode_password(password):
    """解密"""
    encoded_bytes = password.encode('utf-8')
    password_bytes = base64.b64decode(encoded_bytes)
    decoded_str = password_bytes.decode('utf-8')
    return decoded_str[:-4]


def parse_config(server_name):
    """从配置文件加载数据库配置"""
    config = configparser.ConfigParser()
    config.read(f'server/{server_name}-config.ini', encoding='utf-8')
    logging.info(f"服务器： {server_name}")
    return config


port_pool = set(range(12345, 22400))


def get_local_bind_port():
    with threading.Lock():
        if not port_pool:
            raise Exception("No available ports")
        port = port_pool.pop()
        return port


def release_local_bind_port(port):
    with threading.Lock():
        port_pool.add(port)


def connect_to_server(server_config):
    """数据库连接"""
    connection = None  # 初始化数据库连接
    try:
        # if server_config['SslMode'] == 'on':
        #     local_bind_port = get_local_bind_port()
        #     mysql_password = decode_password(server_config.get('password', raw=True))
        #     ssh_password = decode_password(server_config['ssh_password'])
        #     server = SSHTunnelForwarder(
        #         ssh_address_or_host=(server_config['ssh_host'], server_config.getint('ssh_port')),
        #         ssh_username=server_config['ssh_user'],  # 跳转机的用户
        #         ssh_password=ssh_password,
        #         local_bind_address=('127.0.0.1', local_bind_port),
        #         remote_bind_address=(server_config['host'], server_config.getint('port'))
        #     )
        #     server.start()
        #     connection = pymysql.connect(
        #         host='127.0.0.1',
        #         port=local_bind_port,
        #         database=server_config['database'],
        #         user=server_config['user'],
        #         password=mysql_password,
        #         autocommit=True
        #     )
        #
        # else:
        mysql_password = decode_password(server_config.get('password', raw=True))
        connection = pymysql.connect(
            host=server_config['host'],
            port=server_config.getint('port'),
            database=server_config['database'],
            user=server_config['user'],
            password=mysql_password,
            autocommit=True
        )
        logging.info(f"开始操作数据库：{server_config}")
        return connection
    except pymysql.err.OperationalError as err:
        logging.error(f"Error connecting to MySQL Platform: {err}")
        popup_manager.message_signal.emit(f"无法连接到数据库：{err}")
        return None
    except Exception as e:
        logging.error(f"其他错误：{e}")
        popup_manager.message_signal.emit(f"其他错误：{e}")
        return None


def split_statements(sql_text):
    # 正则匹配所有CREATE语句
    create_pattern = r"(CREATE\s+(PROCEDURE|FUNCTION|TRIGGER|EVENT|VIEW)\s+`[a-zA-Z0-9_]+`)"
    matches = list(re.finditer(create_pattern, sql_text))

    if len(matches) == 1:
        return [sql_text]

    statements = []
    start = 0

    for match in matches:
        end = match.start()
        if start != end:
            statements.append(sql_text[start:end].strip())
        start = match.start()

    # 最后一段
    statements.append(sql_text[start:].strip())

    return statements


def clean_sql(sql):
    """规则"""
    print(f"规则：{sql}")
    use_pattern = r"USE\s+`?(\w+)`?\$"
    use_match = re.search(use_pattern, sql, re.IGNORECASE)
    if use_match:
        database_name = use_match.group(1)
        database_name = f"{database_name}."
    else:
        database_name = ""

    sql = re.sub(r'/\*!\d+\s*', "", sql)
    sql = re.sub(r'\*/\s*\$\$', ";", sql)
    sql = re.sub(r'\$\$', ";", sql)
    sql = re.sub(r'\*/\s*;', ";", sql)
    sql = re.sub(r"END\$\$", "END;", sql)
    sql = re.sub(r"END\s\$\$", "END;", sql)
    sql = re.sub(r"DELIMITER\s*\$\$", "", sql)
    sql = re.sub(r"DELIMITER\s*;", "", sql)
    sql = re.sub(r"DELIMITER\s*;", "", sql)
    sql = re.sub(r"USE\s+`[a-zA-Z0-9_]+`;", "", sql)
    sql = re.sub(r"DROP PROCEDURE IF EXISTS `([a-zA-Z0-9_]+)`;", "", sql)
    return sql, database_name

def create_procedure(connection, section, sql, sql_type, database_name, _name, result_df):
    """存储过程建立"""
    create_index = sql.find("CREATE")

    sql = re.sub(rf"CREATE\s+PROCEDURE\s+`{_name}`",
                 f"CREATE PROCEDURE {database_name}`{_name}`", sql, flags=re.IGNORECASE)
    sql = re.sub(rf"CREATE\s+DEFINER=`[^`]+`@`[^`]+`\s+PROCEDURE\s+`{_name}`",
                 f"CREATE PROCEDURE {database_name}`{_name}`", sql, flags=re.IGNORECASE)

    String_sql = ''
    check_proc_exist_sql = ''
    if create_index != -1:
        String_sql = sql[create_index:]
    try:
        cursor = connection.cursor()
        if sql_type == 'PROCEDURE':
            check_proc_exist_sql = f"SHOW CREATE PROCEDURE {database_name}{_name} ;"
        elif sql_type == 'FUNCTION':
            check_proc_exist_sql = f"SHOW CREATE FUNCTION {database_name}{_name} ;"
        elif sql_type == 'TRIGGER':
            check_proc_exist_sql = f"SHOW CREATE TRIGGER {database_name}{_name} ;"
        else:
            pass
        print(String_sql)
        try:
            cursor.execute(check_proc_exist_sql)
            result = cursor.fetchone()
        except pymysql.err.OperationalError as e:
            error_code, error_message = e.args

            if error_code == 1305:  # 不存在
                result = None
            else:
                print("存在的")
                raise e
        if result is None:
            # 不存在，直接创建
            print(f"直接创建：{database_name}{String_sql}")
            cursor.execute(String_sql)
            print("创建完成")
            # connection.commit()
            temp_df = pd.DataFrame(
                {'db_name': [database_name.replace(".", "")], sql_type: [_name], '执行结果': ['成功'],
                 '服务器组': [section]})
            result_df = pd.concat([result_df, temp_df], ignore_index=True)
            print(f"{section}：{database_name}.{_name}：创建成功")
            logging.info(f"{section}：{database_name}.{_name}：创建成功")
        else:
            print(f"{database_name}{_name}已经存在")
            if sql_type == 'PROCEDURE':
                cursor.execute(f"DROP PROCEDURE IF EXISTS {database_name}{_name} ;")
            elif sql_type == 'FUNCTION':
                cursor.execute(f"DROP FUNCTION IF EXISTS {database_name}{_name} ;")
            elif sql_type == 'TRIGGER':
                cursor.execute(f"DROP TRIGGER IF EXISTS {database_name}{_name} ;")
            print(f"{database_name}{_name}已经删除")
            print(f"执行{String_sql}")
            cursor.execute(String_sql)
            connection.commit()
            temp_df = pd.DataFrame(
                {'db_name': [database_name.replace(".", "")], sql_type: [_name], '执行结果': ['成功'],
                 '服务器组': [section]})
            result_df = pd.concat([result_df, temp_df], ignore_index=True)
            print(f"{database_name}{_name}已经更新")
            logging.info(f"{section}：{database_name}{_name}：更新成功")

        cursor.close()
    except pymysql.err.InternalError as e:
        error_code, error_message = e.args
        logging.error(f"{section}：{database_name}{_name}：操作失败: {error_message}")
        popup_manager.message_signal.emit(
            f"{section}：{database_name}{_name}：操作失败: {error_message}")

        temp_df = pd.DataFrame({'执行结果': [f'失败:{str(error_message)}'], '服务器组': [section]})
        result_df = pd.concat([result_df, temp_df], ignore_index=True)


    return result_df


def execute_sql(connection, sql, section, type, db_name, _name, result_df):
    """执行sql语句"""
    try:
        global Process_df
        cursor = connection.cursor()
        Process_id = connection.thread_id()
        row = {'Process_id': Process_id, 'Status': 'Running', '服务器组': section}
        Process_df = Process_df.append(row, ignore_index=True)
        logging.info(f"当前执行进程->{section}:{Process_id}")
        print(f"{datetime.datetime.now()}当前执行进程->{section}:{Process_id}")

        if (sql.strip().upper().startswith('SELECT') or sql.strip().upper().startswith('SHOW')
                or sql.strip().upper().startswith('WITH') or sql.strip().upper().startswith(
                    'DESC') or sql.strip().upper().startswith('EXPLAIN')):
            sql_statements = sqlparse.split(sql)
            for sql_statement in sql_statements:
                logging.info(f"执行SQL语句：{sql_statement}")
                cursor.execute(sql_statement)
            results = cursor.fetchall()
            temp_df = pd.DataFrame(results, columns=[desc[0] for desc in cursor.description])
            temp_df['服务器组'] = section
            result_df = pd.concat([result_df, temp_df], ignore_index=True)
            logging.info(f"{section}：执行成功")
        elif sql.upper().startswith('CALL'):
            cursor.execute(sql)
            if cursor.rowcount == -1:
                # connection.commit()
                temp_df = pd.DataFrame({'执行结果': ['成功'], '服务器组': [section]})
                result_df = pd.concat([result_df, temp_df], ignore_index=True)

            else:
                # connection.commit()
                results = cursor.fetchall()

                if cursor.rowcount == 0 and cursor.description:
                    temp_df = pd.DataFrame(columns=[desc[0] for desc in cursor.description])
                    temp_df['服务器组'] = section
                    result_df = pd.concat([result_df, temp_df], ignore_index=True)
                elif cursor.rowcount == 0 and not cursor.description:
                    temp_df = pd.DataFrame({'执行结果': ['成功'], '服务器组': [section]})
                    result_df = pd.concat([result_df, temp_df], ignore_index=True)
                else:
                    temp_df = pd.DataFrame(results, columns=[desc[0] for desc in cursor.description])
                    temp_df['服务器组'] = section
                    result_df = pd.concat([result_df, temp_df], ignore_index=True)
        elif re.search(r"CREATE\s+(?:FUNCTION|PROCEDURE|VIEW|EVENT|TRIGGER)", sql):
            print("创建存储过程/函数/事件/视图/触发器")
            try:
                result_df = create_procedure(connection, section, sql, type, db_name, _name, None)
            except Exception as e:
                logging.error(f"{section}:{e} ")
        else:
            sql_statements = sqlparse.split(sql)
            print(f"执行中：{sql_statements}")
            success_count = 0
            failure_count = 0
            try:
                for sql_statement in sql_statements:
                    logging.info(f"执行SQL语句：{sql_statement}")
                    cursor.execute(sql_statement)
                    # connection.commit()
                    count = cursor.rowcount
                    if count > 0:
                        success_count += count
                    elif count == 0:
                        success_count += 1
                    else:
                        failure_count += 1
                temp_df = pd.DataFrame(
                    {'成功数': [success_count], '失败数': [failure_count], '服务器组': [section]})
                result_df = pd.concat([result_df, temp_df], ignore_index=True)
                logging.info(f"SQL statements executed successfully!")
            except Exception as e:
                popup_manager.message_signal.emit(f"{section}错误内容: {str(e)}")

    except pymysql.err.OperationalError as e:
        if e.args[0] == 2013:
            logging.error(f"{section} 执行失败: 连接到 MySQL 服务器时连接丢失")

            temp_df = pd.DataFrame({'执行结果': ['执行失败: 连接到 MySQL 服务器时连接丢失'], '服务器组': [section]})
            result_df = pd.concat([result_df, temp_df], ignore_index=True)

        else:
            popup_manager.message_signal.emit(f"{section} 执行失败: {str(e)}")
            logging.error(f"{section} 执行失败: {str(e)}")

            temp_df = pd.DataFrame({'执行结果': [f'失败:{str(e)}'], '服务器组': [section]})
            result_df = pd.concat([result_df, temp_df], ignore_index=True)

    except pymysql.err.ProgrammingError as e:

        popup_manager.message_signal.emit(f"错误内容: {section}:{str(e)}")
        logging.error(f"{section} 执行失败: {str(e)}")

        temp_df = pd.DataFrame({'执行结果': [f'执行失败: {str(e)}'], '服务器组': [section]})
        result_df = pd.concat([result_df, temp_df], ignore_index=True)

    return result_df

def kill_sql(connection, section, result_df):
    """杀进程"""
    matching_processes = Process_df[Process_df['服务器组'] == section]
    if not matching_processes.empty:
        for index, process_to_kill in matching_processes.iterrows():
            process_id_to_kill = process_to_kill['Process_id']
            server_group = process_to_kill['服务器组']
            try:
                with connection.cursor() as kill_cursor:
                    kill_cursor.execute(f"KILL {int(process_id_to_kill)}")
                    logging.info(f"KILL进程ID-> {server_group}:{process_id_to_kill}")

                    temp_df = pd.DataFrame({
                        'Process_id': [process_id_to_kill],
                        'Status': ["stopped"],
                        '服务器组': [server_group]
                    })

                    result_df = pd.concat([result_df, temp_df], ignore_index=True)
            except Exception as e:
                logging.error(f"执行 KILL 命令时发生错误: {str(e)} 进程ID: {process_id_to_kill}")

                temp_df = pd.DataFrame({
                    'Process_id': [process_id_to_kill],
                    'Status': ["failed"],
                    '服务器组': [server_group]
                })
                result_df = pd.concat([result_df, temp_df], ignore_index=True)

        return result_df
    else:
        logging.info(f"没有找到属于 {section} 服务器组的进程ID")
        return result_df

class Ui_MainWindow(object):
    def __init__(self):
        super().__init__()
        self.sqlTextEdit = None
        self.selected_rows = None
        self.selected_colums = None
        self.file_path = ""
        self.selected_rows = []
        self.tab_table_map = {}
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.setWindowTitle("DBtool")
        MainWindow.setWindowIcon(QIcon("resources/icon.ico"))
        MainWindow.setEnabled(True)
        MainWindow.resize(1024, 768)

        menubar = self.menuBar()
        server_menu = menubar.addMenu("菜单")

        add_server_action = QAction("添加服务器", self)
        add_server_action.triggered.connect(self.open_server_dialog)
        server_menu.addAction(add_server_action)

        view_pross_action = QAction("看进程", self)
        view_pross_action.triggered.connect(self.open_process_dialog)
        server_menu.addAction(view_pross_action)

        # copy_table = QAction("将表复制到不同的主机/数据库...", self)
        # copy_table.triggered.connect(self.copy_table_to)
        # server_menu.addAction(copy_table)

        self.centralwidget = QtWidgets.QWidget(parent=MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        layout = QtWidgets.QVBoxLayout(self.centralwidget)
        font = QtGui.QFont("Courier New")
        font.setPointSize(9)
        self.row1Layout = QtWidgets.QHBoxLayout()
        self.sever_1 = QtWidgets.QGroupBox("版本", parent=self.centralwidget)
        self.sever_1_Layout = QtWidgets.QHBoxLayout(self.sever_1)
        self.serverComboBox = QtWidgets.QComboBox(parent=self.centralwidget)
        self.load_server_names()  # 加载服务器列表
        self.serverComboBox.setFont(font)
        self.serverComboBox.currentIndexChanged.connect(lambda: self.updateWindowTitle(MainWindow))
        self.sever_1_Layout.addWidget(self.serverComboBox)
        self.row1Layout.addWidget(self.sever_1)
        self.fast_query = QtWidgets.QGroupBox("快捷查询", parent=self.centralwidget)
        self.fast_query_Layout = QtWidgets.QHBoxLayout(self.fast_query)
        self.CODE = QtWidgets.QLineEdit(parent=self.centralwidget)
        self.CODE.setPlaceholderText("酒店CODE")
        self.CODE.setFont(font)
        self.fast_query_Layout.addWidget(self.CODE)
        self.desc = QtWidgets.QLineEdit(parent=self.centralwidget)
        self.desc.setPlaceholderText("酒店名称")
        self.desc.setFont(font)
        self.fast_query_Layout.addWidget(self.desc)
        self.id = QtWidgets.QLineEdit(parent=self.centralwidget)
        self.id.setPlaceholderText("酒店ID")
        self.id.setFont(font)
        self.fast_query_Layout.addWidget(self.id)
        self.LIMIT = QtWidgets.QLineEdit(parent=self.centralwidget)
        self.LIMIT.setToolTip("限制查询条数")
        self.LIMIT.setFont(font)
        self.LIMIT.setText("100")
        self.fast_query_Layout.addWidget(self.LIMIT)
        self.executeButton = QtWidgets.QPushButton(parent=self.centralwidget)
        self.executeButton.setText("执行 Ctrl+F9")
        self.executeButton.setFont(font)
        self.fast_query_Layout.addWidget(self.executeButton)
        self.executeButton.setShortcut("Ctrl+F9")  # 设置快捷键
        self.executeButton.clicked.connect(self.Find_Wyn_Hotel)
        self.row1Layout.addWidget(self.fast_query)
        layout.addLayout(self.row1Layout)
        self.row2Layout = QtWidgets.QHBoxLayout()

        self.import_box = QtWidgets.QGroupBox("数据导入", parent=self.centralwidget)
        self.import_box_Layout = QtWidgets.QGridLayout(self.import_box)
        self.checkbox = QtWidgets.QCheckBox(parent=self.centralwidget)
        self.checkbox.setText("启用")
        self.checkbox.move(20, 20)
        self.checkbox.stateChanged.connect(self.toggleButtons)
        self.toggleButtons(0)
        self.import_box_Layout.addWidget(self.checkbox, 0, 0)
        self.dbname = QtWidgets.QLineEdit(parent=self.centralwidget)
        self.dbname.setToolTip("库名")
        self.dbname.setText("migrate_db")
        self.dbname.setFont(font)
        self.dbname.setEnabled(False)
        self.import_box_Layout.addWidget(self.dbname, 0, 1)
        self.crate_table = QtWidgets.QLineEdit(parent=self.centralwidget)
        self.crate_table.setPlaceholderText("创建表名")
        self.crate_table.setToolTip("默认Excel表单名")
        self.crate_table.setFont(font)
        self.crate_table.setEnabled(False)
        self.import_box_Layout.addWidget(self.crate_table, 0, 2)
        self.select_button = QPushButton(parent=self.centralwidget)
        self.select_button.setText("Excel文件")
        self.select_button.clicked.connect(self.select_excel)
        self.select_button.setEnabled(False)
        self.import_box_Layout.addWidget(self.select_button, 1, 0)
        self.sheet_dropdown = QComboBox()
        self.sheet_dropdown.setFont(font)
        self.sheet_dropdown.setEnabled(False)
        self.import_box_Layout.addWidget(self.sheet_dropdown, 1, 1)
        self.import_button = QPushButton(parent=self.centralwidget)
        self.import_button.setText("导入")
        self.import_button.setFont(font)
        self.import_button.clicked.connect(self.import_data)
        self.import_button.setEnabled(False)
        self.import_box_Layout.addWidget(self.import_button, 1, 2)
        self.row2Layout.addWidget(self.import_box)
        self.sql_box = QtWidgets.QGroupBox("搞SQL脚本", parent=self.centralwidget)
        self.sql_box_Layout = QtWidgets.QHBoxLayout(self.sql_box)
        self.sql_file = QPushButton(parent=self.centralwidget)
        self.sql_file.setText("文件夹")
        self.sql_file.clicked.connect(self.select_file)
        self.sql_file.setEnabled(True)
        self.sql_box_Layout.addWidget(self.sql_file)
        self.sqlButton2 = QtWidgets.QPushButton(parent=self.centralwidget)
        self.sqlButton2.setText("搞脚本")
        # self.sqlButton2.setFixedSize(130, 32)
        self.sqlButton2.setFont(font)
        self.sqlButton2.clicked.connect(self.execute_sql_scripts)
        self.sql_box_Layout.addWidget(self.sqlButton2)
        self.row2Layout.addWidget(self.sql_box)
        self.timeLabel = QLabel()
        self.timeLabel.setFont(font)
        self.end_box = QtWidgets.QGroupBox("开搞", parent=self.centralwidget)
        self.end_box_Layout = QtWidgets.QGridLayout(self.end_box)
        self.executeButton2 = QtWidgets.QPushButton(parent=self.centralwidget)
        self.executeButton2.setText("开搞 F9")
        # self.executeButton2.setFixedSize(130, 47)
        self.executeButton2.setFont(font)
        self.executeButton2.setShortcut("F9")  # 设置快捷键
        self.executeButton2.clicked.connect(self.execute_button_clicked)
        self.end_box_Layout.addWidget(self.executeButton2, 0, 0)
        self.executeButton3 = QtWidgets.QPushButton(parent=self.centralwidget)
        self.executeButton3.setText("不搞了 F12")
        # self.executeButton3.setFixedSize(130, 47)
        self.executeButton3.setFont(font)
        self.executeButton3.setShortcut("F12")  # 设置快捷键
        self.executeButton3.setEnabled(False)
        self.executeButton3.clicked.connect(self.stop_button_clicked)
        self.end_box_Layout.addWidget(self.executeButton3, 1, 0)

        self.execute_PMS = QtWidgets.QCheckBox(parent=self.centralwidget)
        self.execute_PMS.setText("门店")
        self.execute_PMS.move(20, 20)
        self.execute_PMS.setChecked(True)
        self.execute_PMS.setEnabled(False)
        self.execute_group = QtWidgets.QCheckBox(parent=self.centralwidget)
        self.execute_group.setText("集团")
        self.execute_group.move(20, 20)
        self.execute_member = QtWidgets.QCheckBox(parent=self.centralwidget)
        self.execute_member.setText("会员")
        self.execute_member.move(20, 20)
        self.end_box_Layout.addWidget(self.execute_PMS, 0, 1)
        self.end_box_Layout.addWidget(self.execute_group, 1, 1)
        self.end_box_Layout.addWidget(self.execute_member, 2, 1)
        self.row2Layout.addWidget(self.end_box)

        layout.addLayout(self.row2Layout)
        splitter = QSplitter(Qt.Vertical)
        self.sqlTextEdit = QtWidgets.QTextEdit(parent=self.centralwidget)
        self.sqlTextEdit.setFont(font)
        self.sqlTextEdit.setText("select * from hotel limit 1;")
        self.sqlTextEdit.setStyleSheet("""  
                    QTextEdit {  
                        background-color: #434343;  /* 深色背景 */  
                        color: #d4d4d4;            /* 浅灰色文字 */  
                        selection-background-color: #3a3a3a; /* 选中时的背景色 */  
                        selection-color: #ffffff;    /* 选中时的文字色 */  
                        border: 1px solid #444444;   /* 边框颜色 */  
                        border-radius: 2px;         /* 边框圆角 */  
                    }  
                """)
        splitter.addWidget(self.sqlTextEdit)

        self.tab_widget = QtWidgets.QTabWidget(self.centralwidget)
        self.tab_widget.tabBar().setTabsClosable(True)
        self.tab_widget.tabBar().tabCloseRequested.connect(self.close_tab)
        self.tab_widget.tabBar().setMouseTracking(True)

        self.tab_widget.currentChanged.connect(self.setup_table_context_menu)
        self.context_menu = QMenu(self.centralwidget)
        self.action_export = QAction("导出为excel", self.centralwidget)
        self.action_export_csv = QAction("导出为CSV", self.centralwidget)
        self.filter = QAction("查找", self.centralwidget)

        self.action_export.triggered.connect(self.export_to_excel)
        self.action_export_csv.triggered.connect(self.export_to_csv)
        self.filter.triggered.connect(self.filterTable)

        self.context_menu.addAction(self.action_export)
        self.context_menu.addAction(self.action_export_csv)
        self.context_menu.addAction(self.filter)

        splitter.addWidget(self.sqlTextEdit)
        splitter.addWidget(self.tab_widget)
        layout.addWidget(splitter)

        layout = QtWidgets.QVBoxLayout(self.centralwidget)
        layout.addWidget(self.tab_widget)

        MainWindow.setCentralWidget(self.centralwidget)
        self.retranslateUi(MainWindow)

        self.timer = QTimer(self)
        self.timer.timeout.connect(lambda: self.updateRuntime(MainWindow))
        # self.timer.start(10)  # 每 10 毫秒更新一次

        # 标志变量来判断是否需要更新时间
        self.isRunning = False
        self.elapsed_time = 0  # 记录已经经过的时间（单位为毫秒）


        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def open_server_dialog(self):
        dialog = ServerDialog()
        if dialog.exec_() == QDialog.Accepted:
            # 获取用户输入的数据
            service_group_name = dialog.service_group_name.text()
            server_data = {
                "host": dialog.host.text(),
                "port": dialog.port.text(),
                "user": dialog.user.text(),
                "password": encode_password(f"{dialog.password.text()}zwy"),
                "database": dialog.database.text(),
                "ssh_host": dialog.ssh_host.text() if dialog.ssh_checkbox.isChecked() else None,
                "ssh_port": dialog.ssh_port.text() if dialog.ssh_checkbox.isChecked() else None,
                "ssh_user": dialog.ssh_user.text() if dialog.ssh_checkbox.isChecked() else None,
                "ssh_password": encode_password(
                    f"{dialog.ssh_password.text()}") if dialog.ssh_checkbox.isChecked() else None,
                "SslMode": "on" if dialog.ssh_checkbox.isChecked() else "no"  # 根据需要设置SslMode
            }
            ini_file_path = f"server/{self.serverComboBox.currentText().strip()}-config.ini"  # 替换为您的INI文件路径
            config = configparser.ConfigParser()

            # 检查是否存在INI文件，如果不存在，则创建一个
            if os.path.exists(ini_file_path):
                config.read(ini_file_path, encoding='utf-8')
            else:
                config['服务器'] = {}

            # 添加新的项目部服务器
            config[service_group_name] = {k: v for k, v in server_data.items() if v is not None}

            # 写入到INI文件
            with open(ini_file_path, 'w', encoding='utf-8') as configfile:
                config.write(configfile)

            print(f"{service_group_name} 的服务器信息已写入到 {ini_file_path}")

    def open_process_dialog(self):
        server_name = self.serverComboBox.currentText().strip()
        execute_group = self.execute_group
        execute_member = self.execute_member
        dialog = ProcessDialog(server_name, execute_member, execute_group, self)
        dialog.exec_()

    def copy_table_to(self):
        server_name = self.serverComboBox.currentText().strip()
        execute_group = self.execute_group
        execute_member = self.execute_member
        dialog = CopyTableto_Dialog(server_name, execute_member, execute_group, self)
        dialog.exec_()

    def closeEvent(self, event):
        reply = QtWidgets.QMessageBox.question(self, '确认退出',
                                               '您确定要退出吗？', QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                                               QtWidgets.QMessageBox.No)

        if reply == QtWidgets.QMessageBox.Yes:
            event.accept()  # Close the application
        else:
            event.ignore()  # Ignore the close event

    def close_tab(self, index):
        self.tab_widget.removeTab(index)

    def setup_table_context_menu(self):
        # 设置每个表格的右键菜单
        for i in range(self.tab_widget.count()):
            table = self.tab_widget.widget(i)
            table.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
            table.customContextMenuRequested.connect(self.show_context_menu)

    def show_context_menu(self, pos):
        self.context_menu.exec_(self.tab_widget.mapToGlobal(pos))

    def execute_button_clicked(self):
        server_name = self.serverComboBox.currentText().strip()
        sql = self.sqlTextEdit.toPlainText().strip()
        timeLabel = self.timeLabel
        execute_group = self.execute_group
        execute_pms = self.execute_PMS
        execute_member = self.execute_member
        self.startExecution()
        self.thread = Thread_1(server_name, sql, timeLabel, execute_group, execute_pms, execute_member)
        self.thread.start()
        self.thread.result_ready.connect(self.on_result_ready)
        self.serverComboBox.setEnabled(False)
        self.executeButton2.setEnabled(False)
        self.executeButton3.setEnabled(True)

    def stop_button_clicked(self):
        server_name = self.serverComboBox.currentText().strip()
        timeLabel = self.timeLabel
        execute_group = self.execute_group
        execute_pms = self.execute_PMS
        execute_member = self.execute_member
        self.thread = Thread_2(server_name, timeLabel, execute_group, execute_pms, execute_member)
        self.thread.start()
        self.thread.stop_click.connect(self.stop_click)
        self.executeButton2.setEnabled(True)
        self.executeButton3.setEnabled(False)

    def filterTable(self):
        try:
            current_index = self.tab_widget.currentIndex()  # 获取当前选项卡索引
            current_tab = self.tab_widget.widget(current_index)  # 获取当前选项卡

            text, okPressed = QtWidgets.QInputDialog.getText(current_tab, "查找", "输入查找文本:",
                                                             QtWidgets.QLineEdit.Normal, "")
            if okPressed and text.strip():
                selection_model = current_tab.selectionModel()
                model = current_tab.model()

                selected_indexes = []
                self.matched_rows = []
                for row in range(model.rowCount()):
                    for column in range(model.columnCount()):
                        index = model.index(row, column)
                        if index.isValid() and text.lower() in index.data().lower():  # 模糊搜索，忽略大小写
                            selected_indexes.append(index)
                            if row not in self.matched_rows:
                                self.matched_rows.append(row)
                selection_model.clearSelection()
                if selected_indexes:
                    new_selection = QItemSelection()
                    for index in selected_indexes:
                        new_selection.select(index, index)

                    selection_model.select(new_selection, QItemSelectionModel.ClearAndSelect)
                    view = current_tab  # 假设 current_tab 是 QTableView
                    first_matched_index = selected_indexes[0]
                    view.scrollTo(first_matched_index, QAbstractItemView.PositionAtTop)
                    self.current_matched_row_index = 0
        except Exception as e:
            print(e)

    def deleteTable(self):
        current_index = self.tab_widget.currentIndex()
        if current_index != -1:  # 确保当前选项卡有效
            self.tab_widget.removeTab(current_index)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        server = self.serverComboBox.currentText()
        MainWindow.setWindowTitle(_translate("MainWindow", f"DB_Tool-{server} "))


    def updateWindowTitle(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        server = self.serverComboBox.currentText()
        MainWindow.setWindowTitle(_translate("MainWindow", f"DB_Tool-{server}"))
    def startExecution(self):
        """模拟开始执行的过程"""
        self.isRunning = True
        self.elapsed_time = 0
        self.timer.start(10)

    def finishExecution(self,MainWindow):
        """模拟执行完毕的操作"""
        _translate = QtCore.QCoreApplication.translate
        self.isRunning = False
        self.timer.stop()
        self.elapsed_time += 10
        # 转换成分钟:秒:毫秒 格式
        minutes = self.elapsed_time // 60000
        seconds = (self.elapsed_time % 60000) // 1000
        milliseconds = self.elapsed_time % 1000
        # 格式化时间为 "mm:ss:SSS"
        formatted_time = f"{minutes:02}:{seconds:02}:{milliseconds:03}"
        MainWindow.setWindowTitle(_translate("MainWindow", f"DB_Tool- 耗时:{formatted_time} - 完成！"))

    def dorp_table(self,MainWindow):
        """模拟执行完毕的操作"""
        _translate = QtCore.QCoreApplication.translate
        self.isRunning = False
        self.timer.stop()
        MainWindow.setWindowTitle(_translate("MainWindow", f"DB_Tool-渲染表格中"))

    def updateRuntime(self,MainWindow):
        """更新窗口标题，显示时间"""
        _translate = QtCore.QCoreApplication.translate
        if self.isRunning:
            self.elapsed_time += 10
            # 转换成分钟:秒:毫秒 格式
            minutes = self.elapsed_time // 60000
            seconds = (self.elapsed_time % 60000) // 1000
            milliseconds = self.elapsed_time % 1000
            # 格式化时间为 "mm:ss:SSS"
            formatted_time = f"{minutes:02}:{seconds:02}:{milliseconds:03}"
            # 更新窗口标题
            MainWindow.setWindowTitle(f"DBTool - 耗时:{formatted_time}")

    def toggleButtons(self, state):
        try:
            if state == 2:  # 2表示选中状态
                self.dbname.setEnabled(True)
                self.crate_table.setEnabled(True)
                self.select_button.setEnabled(True)
                self.sheet_dropdown.setEnabled(True)
                self.import_button.setEnabled(True)
            elif state == 0:
                self.dbname.setEnabled(False)
                self.crate_table.setEnabled(False)
                self.select_button.setEnabled(False)
                self.sheet_dropdown.setEnabled(False)
                self.import_button.setEnabled(False)
        except Exception as e:
            pass

    def load_server_names(self):
        config_files = [f for f in os.listdir('server') if f.endswith('-config.ini')]
        server_names = [f.replace('-config.ini', '') for f in config_files]
        self.serverComboBox.addItems(server_names)

    def Find_Wyn_Hotel(self):
        hotelcode = self.CODE.text()
        hotelid = self.id.text()
        desc = self.desc.text()
        hotellimit = self.LIMIT.text()
        server_name = self.serverComboBox.currentText().strip()

        logging.info(f"当前执行集团: {server_name}")
        config = parse_config(server_name)
        for section in config.sections():
            if section == 'group':
                try:
                    server = config[section]
                    conn = connect_to_server(server)
                    if conn is not None:
                        try:
                            with conn.cursor() as cursor:
                                sql = "SELECT hotel_group_id,id,CODE,sta,audit,descript,descript_en,descript_short,country,city,address1,address2,phone,fax,phone_rsv,website,remark,create_user,create_datetime,modify_user,modify_datetime,province_code,city_code,district_code,brand_code,score,manage_type,client_type,online_check,server_name FROM hotel "
                                conditions = []
                                if hotelcode:
                                    conditions.append(f"CODE = '{hotelcode}'")
                                if hotelid:
                                    conditions.append(f"ID = {hotelid}")
                                if desc:
                                    conditions.append(f"descript LIKE '%{desc}%'")
                                if conditions:
                                    sql += " WHERE " + " AND ".join(conditions)
                                sql += " ORDER BY id"
                                if hotellimit:
                                    sql += f" LIMIT {hotellimit};"
                                logging.info(f"执行的sql: {sql}")
                                print(sql)
                                start_time = datetime.datetime.now()
                                cursor.execute(sql)
                                elapsed_time = datetime.datetime.now() - start_time
                                formatted_time = elapsed_time.total_seconds()
                                print("查询了吗")
                                results = cursor.fetchall()
                                column = [desc[0] for desc in cursor.description]
                                self.dorp_tablenew(column, results, "快捷查询")
                        except Exception as e:
                            popup_manager.message_signal.emit(f"错误内容: {str(e)}")
                            logging.error(f"发生异常: {e}")
                        finally:
                            # 关闭数据库连接
                            conn.close()

                except  Exception as e:
                    QMessageBox.critical(None, "连接错误", f"无法连接到数据库：{e}")
                    logging.error(f"无法连接到数据库：{e}")
                    return

    def select_excel(self):
        try:
            self.sheet_dropdown.clear()
            self.file_path, _ = QFileDialog.getOpenFileName(self.centralwidget, "选择Excel文件", "",
                                                            "Excel files (*.xlsx *.xls)")
            print(f"cd:/{self.file_path}")
            if self.file_path == "":
                self.select_button.setText("Excel文件")
                self.select_button.setToolTip("")

            if self.file_path != "":
                self.select_button.setText("已选择")
                self.select_button.setToolTip(self.file_path)

            if self.file_path:
                xl = pd.ExcelFile(self.file_path)
                sheet_names = xl.sheet_names
                self.sheet_dropdown.addItems(sheet_names)
        except Exception as e:
            print(f"Error occurred: {str(e)}")

    def export_to_excel(self):
        try:
            current_index = self.tab_widget.currentIndex()
            print(current_index)
            # 获取当前标签页中的表格
            current_tab = self.tab_widget.widget(current_index)
            model = current_tab.model()  # 假设每个标签页都有一个模型
            if model is None or model.columnCount() == 0:
                QMessageBox.information(self, "！", "无数据")
                return
            file_dialog = QFileDialog()
            file_dialog.setDefaultSuffix("xlsx")
            file_path, _ = file_dialog.getSaveFileName(None, "保存文件", "", "Excel Files (*.xlsx)")
            if not file_path:
                return
            try:
                headers = [model.headerData(column, Qt.Horizontal) for column in range(model.columnCount())]
                data = []
                for row in range(model.rowCount()):
                    rowData = []
                    for column in range(model.columnCount()):
                        index = model.index(row, column)
                        rowData.append(model.data(index))
                    data.append(rowData)
            except Exception as e:
                print(e)
            # 如果有选中行，则只导出选中行的数据
            selection_model = current_tab.selectionModel()
            selected_indexes = selection_model.selectedRows()
            selected_rows = [index.row() for index in selected_indexes] if selected_indexes else []
            if selected_rows:
                data = [data[row] for row in selected_rows]
            export_thread = ExportThread(data, headers, file_path)
            print(datetime.datetime.now(), "------------开始导出-----------------")
            logging.info(f"{datetime.datetime.now()} ------------开始导出-----------------")
            export_thread.start()
        except Exception as e:
            print(f"Error occurred while exporting to Excel: {str(e)}")

    def export_to_csv(self):
        try:
            current_index = self.tab_widget.currentIndex()
            print(current_index)
            # 获取当前标签页中的表格
            current_tab = self.tab_widget.widget(current_index)
            model = current_tab.model()  # 假设每个标签页都有一个模型
            if model is None or model.columnCount() == 0:
                QMessageBox.information(self, "！", "无数据")
                return

            # 打开文件保存对话框，设置默认文件扩展名为 .csv
            file_dialog = QFileDialog()
            file_dialog.setDefaultSuffix("csv")
            file_path, _ = file_dialog.getSaveFileName(None, "保存文件", "", "CSV Files (*.csv)")
            if not file_path:
                return

            try:
                # 获取表头数据
                headers = [model.headerData(column, Qt.Horizontal) for column in range(model.columnCount())]
                data = []
                for row in range(model.rowCount()):
                    rowData = []
                    for column in range(model.columnCount()):
                        index = model.index(row, column)
                        rowData.append(model.data(index))
                    data.append(rowData)
            except Exception as e:
                print(e)

            # 如果有选中行，则只导出选中行的数据
            selection_model = current_tab.selectionModel()
            selected_indexes = selection_model.selectedRows()
            selected_rows = [index.row() for index in selected_indexes] if selected_indexes else []
            if selected_rows:
                data = [data[row] for row in selected_rows]
            export_thread = ExportThreadCsv(data, headers, file_path)
            print(datetime.datetime.now(), "------------开始导出-----------------")
            logging.info(f"{datetime.datetime.now()} ------------开始导出-----------------")
            export_thread.start()
        except Exception as e:
            print(f"Error occurred while exporting to CSV: {str(e)}")


    def on_export_finished(self):
        QMessageBox.information(self, "完成", "导出完成!")

    def on_result_ready(self,result_df, sql_query):
        # 处理DataFrame并更新UI
        cursor = pd.DataFrame(result_df.columns).reset_index().values.tolist()
        description = [column[1] for column in cursor]
        data = result_df.values.tolist()
        data_tuple = tuple(map(tuple, data))
        # 处理 sql_query
        sql_query = sql_query.replace('\n', ' ').replace('\r', ' ')
        if len(sql_query) > 20:
            sql_query_display = sql_query[:20] + '...'
        else:
            sql_query_display = sql_query
        tab_name = f"{sql_query_display}"
        # self.dorp_table(self)
        self.dorp_tablenew(description, data_tuple, tab_name)
        Process_df.drop(Process_df.index, inplace=True)
        self.executeButton2.setEnabled(True)
        self.executeButton3.setEnabled(False)
        self.serverComboBox.setEnabled(True)
        self.finishExecution(self)

    def stop_click(self, result_df):
        # 处理DataFrame并更新UI
        cursor = pd.DataFrame(result_df.columns).reset_index().values.tolist()
        description = [column[1] for column in cursor]
        data = result_df.values.tolist()
        data_tuple = tuple(map(tuple, data))
        tab_count = self.tab_widget.count()
        tab_name = f"停止结果 {tab_count + 1}"
        Process_df.drop(Process_df.index, inplace=True)
        self.dorp_tablenew(description, data_tuple, tab_name)
        self.executeButton2.setEnabled(True)
        self.serverComboBox.setEnabled(True)
        self.finishExecution(self)

    def select_file(self):
        try:
            self.folder_path = QtWidgets.QFileDialog.getExistingDirectory(self.centralwidget, "选择文件夹")
            print(self.folder_path)
            self.sql_file.setText("已选择")
            self.sql_file.setToolTip(self.folder_path)
        except Exception as e:
            logging.error(f"Error occurred: {str(e)}")
            popup_manager.message_signal.emit(f"Error occurred: {str(e)}")
            print(f"Error occurred: {str(e)}")

    def execute_sql_scripts(self):
        if not hasattr(self, 'folder_path') or not self.folder_path:
            print("请先选择文件夹")
            popup_manager.message_signal.emit("请先选择文件夹")
            return

        result_queue = queue.Queue()  # 创建一个队列来存储每个线程的执行结果
        self.serverComboBox.setEnabled(False)
        try:
            server_name = self.serverComboBox.currentText().strip()
            config = parse_config(server_name)

            def process_server_group(server, section, result_queue):
                connection = connect_to_server(server)
                for filename in os.listdir(self.folder_path):
                    if filename.endswith(".sql"):
                        file_path = os.path.join(self.folder_path, filename)

                        with open(file_path, 'r' ,encoding= 'utf-8' ,errors='ignore' ) as file:
                            concat = file.read()
                            decoded_content = concat.replace('\xa0', ' ')
                            try:
                                sql_query, db_name = clean_sql(decoded_content)
                                print(sql_query)
                                statements = split_statements(sql_query)
                                for sql_segment in statements:
                                    # print(f"SQL execute_sql_scripts 语句: {sql_segment}\n")
                                    if re.search(r"CREATE\s+(?:FUNCTION|PROCEDURE|VIEW|EVENT|TRIGGER)", sql_segment):
                                        # 提取
                                        create_pattern = r"CREATE\s+(?:/[*!].*?[*]/\s*)?(?:DEFINER\s*=\s*`[^`]+`@`[^`]+`\s*)?(TRIGGER|PROCEDURE|FUNCTION|VIEW|EVENT)\s+`([a-zA-Z0-9_]+)`"
                                        match = re.search(create_pattern, sql_segment)
                                        if match:
                                            _type = match.group(1)  # 类型
                                            _name = match.group(2)  # 名称

                                            result = execute_sql(connection, sql_segment, section, _type, db_name, _name,
                                                             None)

                                        if result is not None:
                                            result_queue.put((sql_segment, result))
                                    else:
                                        for statement in sqlparse.parse(sql_segment):
                                            sql_str = str(statement)
                                            logging.info(f"执行SQL语句：{sql_str}")
                                            result = execute_sql(connection, sql_str, section, None, None, None)
                                            if result is not None:
                                                result_queue.put((sql_str, result))
                            except Exception as e:
                                logging.error(f"未完成执行的错误?：=>>  {server}-{e}")
                                popup_manager.message_signal.emit(f"未完成执行的错误?：=>>  {server}-{e}")

            threads = []

            for section in config.sections():
                if section == 'group' and not self.execute_group.isChecked():
                    continue
                if section == 'member' and not self.execute_member.isChecked():
                    continue
                server = config[section]
                thread = threading.Thread(target=process_server_group, args=(server, section, result_queue),
                                          name=f"{section}-Thread")
                thread.start()
                threads.append((thread, server, section))  # 记录线程，服务器和节的元组

            # 等待所有线程完成

            for thread, server, section in threads:
                thread.join()

                logging.info(f"{thread.name} 已执行!")
                print(f"{thread.name} 已执行！！")

            # 在界面上显示结果
            result_df = pd.DataFrame()
            while not result_queue.empty():
                sql_str, result = result_queue.get()
                if result_df.empty:
                    result_df = result
                else:
                    result_df = result_df.append(result, ignore_index=True)

            cursor_df = pd.DataFrame(result_df.columns).reset_index().values.tolist()
            description = [(column[1]) for column in cursor_df]
            data = result_df.values.tolist()

            data_tuple = tuple(map(tuple, data))
            self.dorp_tablenew(description, data_tuple, "脚本结果")
            self.serverComboBox.setEnabled(True)

        except Exception as e:
            print(f"Error occurred: {str(e)}")
            logging.error(f"Error occurred: {str(e)}")
            popup_manager.message_signal.emit(f"Error occurred: {str(e)}")

    def import_data(self):
        cursor_pms = None
        try:
            selected_sheet = self.sheet_dropdown.currentText()
            data = pd.read_excel(self.file_path, sheet_name=selected_sheet, dtype=str)
            server_name = self.serverComboBox.currentText().strip()
            config = parse_config(server_name)
            self.database_name = self.dbname.text()
            self.table_name = self.crate_table.text()
            if self.table_name == '':
                self.table_name = selected_sheet

            for section in config.sections():
                if section == 'group' and not self.execute_group.isChecked():
                    continue
                if section == 'member' and not self.execute_member.isChecked():
                    continue
                server = config[section]
                connection = connect_to_server(server)
                cursor_pms = connection.cursor()
                self.process_data(cursor_pms, data, self.database_name, self.table_name, section)
                connection.commit()
                connection.close()

            QMessageBox.information(None, "成功", f"已成功导入到 {self.database_name} 库，表名：{self.table_name}")
        except Exception as e:
            print(f"导入数据时出现异常: {e}")
            logging.error(f"导入数据异常: {e}")
            QMessageBox.information(None, "错误", f"{e}")
        finally:
            cursor_pms.close()

    def process_data(self, cursor_pms, data, database_name, table_name, section):
        create_table_query = f"CREATE TABLE IF NOT EXISTS {database_name}.{table_name} ("
        for column in data.columns:
            create_table_query += f"`{column}` VARCHAR(255), "
        create_table_query = create_table_query[:-2] + ");"
        cursor_pms.execute(create_table_query)

        data = data.where(pd.notnull(data), None)

        insert_query = f"INSERT INTO {database_name}.{table_name} ({', '.join(data.columns)}) VALUES ({', '.join(['%s' for _ in range(len(data.columns))])})"
        values = [tuple(row) for row in data.values]

        cursor_pms.executemany(insert_query, values)

    def dorp_tablenew(self, column, results, tablename):
        """创建或更新表格"""
        resultTable = QtWidgets.QTableView(self.centralwidget)
        resultTable.setUpdatesEnabled(True)
        resultTable.setSortingEnabled(True)
        resultTable.setStyleSheet("""
            QTableView {
                selection-background-color: #3a3a3a;
                selection-color: #ffffff;
                border: 1px solid #444444;
                border-radius: 2px;
            }
            QTableView::item {
                height: 30px;
            }
        """)
        resultTable.setMinimumHeight(150)
        tab_name = f"{tablename}"

        model = LargeTableModel(results, column)
        proxy_model = QtCore.QSortFilterProxyModel()
        proxy_model.setSourceModel(model)
        resultTable.setModel(proxy_model)

        self.tab_widget.addTab(resultTable, tab_name)
        self.tab_table_map[tablename] = resultTable
        resultTable.resizeColumnsToContents()
        resultTable.resizeRowsToContents()

        last_index = self.tab_widget.count() - 1
        self.tab_widget.setCurrentIndex(last_index)

class ExportThread(threading.Thread):
    def __init__(self, data, headers, file_path, chunk_size=50000, num_threads=10):
        threading.Thread.__init__(self)
        self.data = data
        self.headers = headers
        self.file_path = file_path
        self.chunk_size = chunk_size
        self.num_threads = num_threads
        self.export_finished = False  # 导出完成标志

    def run(self):
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        loop.run_until_complete(self.write_to_excel())

    async def write_to_excel(self):

        workbook = openpyxl.Workbook()

        worksheet = workbook.active
        StartTime = datetime.datetime.now()

        for col_idx, header in enumerate(self.headers, start=1):
            worksheet.cell(row=1, column=col_idx, value=str(header))

        with ThreadPoolExecutor(max_workers=self.num_threads) as executor:
            futures = []
            for chunk_start in range(0, len(self.data), self.chunk_size):
                chunk_end = min(chunk_start + self.chunk_size, len(self.data))
                chunk_data = self.data[chunk_start:chunk_end]
                future = executor.submit(self.write_chunk_to_excel, worksheet, chunk_data, chunk_start + 2)
                futures.append(future)

            await asyncio.gather(*[asyncio.wrap_future(f) for f in futures])

            workbook.save(self.file_path)
            workbook.close()
            EndTime = datetime.datetime.now()
            DiffTime = EndTime - StartTime

            logging.info(f"{datetime.datetime.now()} ------------导出完成-----------------")
            logging.info(f"导出完成,耗时：{round(DiffTime.total_seconds(), 4)}秒")
            print(f"{datetime.datetime.now()} ------------导出完成-----------------")
            print(f"导出完成,耗时：{round(DiffTime.total_seconds(), 4)}秒")
            popup_manager.message_info.emit(f"导出完成,耗时：{round(DiffTime.total_seconds(), 4)}秒")
            self.export_finished = True
    def write_chunk_to_excel(self, worksheet, chunk_data, start_row):
        for row_idx, row_data in enumerate(chunk_data, start=start_row):
            for col_idx, cell_data in enumerate(row_data, start=1):
                worksheet.cell(row=row_idx, column=col_idx, value=str(cell_data))

class ExportThreadCsv(threading.Thread):
    def __init__(self, data, headers, file_path, chunk_size=50000, num_threads=10):
        threading.Thread.__init__(self)
        self.data = data
        self.headers = headers
        self.file_path = file_path
        self.chunk_size = chunk_size
        self.num_threads = num_threads
        self.export_finished = False  # 导出完成标志

    def run(self):
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        loop.run_until_complete(self.write_to_csv())

    async def write_to_csv(self):
        StartTime = datetime.datetime.now()
        logging.info(f"{datetime.datetime.now()} ------------开始导出 CSV -----------------")
        print(f"{datetime.datetime.now()} ------------开始导出 CSV -----------------")

        # 打开文件进行写入
        with open(self.file_path, mode='w', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            # 写入表头
            writer.writerow(self.headers)
            # 准备线程池
            with ThreadPoolExecutor(max_workers=self.num_threads) as executor:
                futures = []
                for chunk_start in range(0, len(self.data), self.chunk_size):
                    chunk_end = min(chunk_start + self.chunk_size, len(self.data))
                    chunk_data = self.data[chunk_start:chunk_end]
                    future = executor.submit(self.write_chunk_to_csv, writer, chunk_data)
                    futures.append(future)

                # 等待所有线程完成
                await asyncio.gather(*[asyncio.wrap_future(f) for f in futures])

        EndTime = datetime.datetime.now()
        DiffTime = EndTime - StartTime
        logging.info(f"{datetime.datetime.now()} ------------导出完成-----------------")
        logging.info(f"导出完成, 耗时：{round(DiffTime.total_seconds(), 4)}秒")
        print(f"{datetime.datetime.now()} ------------导出完成-----------------")
        print(f"导出完成, 耗时：{round(DiffTime.total_seconds(), 4)}秒")

        # 通知导出完成
        popup_manager.message_info.emit(f"导出完成, 耗时：{round(DiffTime.total_seconds(), 4)}秒")
        self.export_finished = True

    def write_chunk_to_csv(self, writer, chunk_data):
        # 将数据写入 CSV 文件
        for row_data in chunk_data:
            writer.writerow(row_data)
class LargeTableModel(QtCore.QAbstractTableModel):
    def __init__(self, data, headers, parent=None):
        super(LargeTableModel, self).__init__(parent)
        self._data = data if isinstance(data, list) else list(data)
        self._headers = headers
        self._sort_order = QtCore.Qt.AscendingOrder
        self._sort_column = -1

    def rowCount(self, parent=QtCore.QModelIndex()):
        return len(self._data)

    def columnCount(self, parent=QtCore.QModelIndex()):
        return len(self._headers)

    def data(self, index, role=QtCore.Qt.DisplayRole):
        if not index.isValid():
            return None

        if role == QtCore.Qt.DisplayRole:
            return str(self._data[index.row()][index.column()])
        return None

    def headerData(self, section, orientation, role=QtCore.Qt.DisplayRole):
        if role != QtCore.Qt.DisplayRole:
            return None

        if orientation == QtCore.Qt.Horizontal:
            return self._headers[section]
        elif orientation == QtCore.Qt.Vertical:
            return str(section + 1)
        return None

    def sort(self, column, order):
        self.layoutAboutToBeChanged.emit()
        self._data.sort(key=lambda row: row[column], reverse=(order == QtCore.Qt.DescendingOrder))
        self.layoutChanged.emit()

    def appendRow(self, row_data):
        """向模型中添加一行数据"""
        if isinstance(row_data, tuple):
            row_data = list(row_data)
        elif not isinstance(row_data, list):
            raise ValueError("row_data must be a list or tuple")

        self.beginInsertRows(QtCore.QModelIndex(), self.rowCount(), self.rowCount())
        self._data.append(row_data)
        self.endInsertRows()




class Thread_1(QThread):
    signal = pyqtSignal()
    result_ready = pyqtSignal(pd.DataFrame, str)

    def __init__(self, server_name, sql, timeLabel, execute_group, execute_pms, execute_member):
        super(Thread_1, self).__init__()
        self.timeLabel = timeLabel
        self.server_name = server_name
        self.sql = sql
        self.result_queue = queue.Queue()
        self.execute_group = execute_group
        self.execute_pms = execute_pms
        self.execute_member = execute_member

    def run(self):
        try:
            server_name = self.server_name
            sql = self.sql
            result_dict = {}
            config = parse_config(server_name)
            start_time = datetime.datetime.now()  # 记录开始时间

            def execute_query(server, section, sql, type, name, db_name):
                connection = connect_to_server(server)
                if connection:
                    try:
                        logging.info(f"执行SQL语句：{sql}")
                        result = execute_sql(connection, sql, section, type, db_name, name, None)
                        if result is not None:
                            self.result_queue.put((sql, result))
                    except Exception as e:
                        logging.error(f"未完成执行的错误：{section}-{e}")
                        popup_manager.message_signal.emit(f"未完成执行的错误：{section}-{e}")
                    finally:
                        connection.close()

            threads = []
            sql_query, db_name = clean_sql(sql)
            # 分段
            statements = split_statements(sql_query)
            for sql_segment in statements:
                for section in config.sections():
                    if section == 'group' and not self.execute_group.isChecked():
                        continue
                    if section == 'member' and not self.execute_member.isChecked():
                        continue
                    server = config[section]
                    # 为每个服务组创建线程
                    if re.search(r"CREATE\s+(?:FUNCTION|PROCEDURE|VIEW|EVENT|TRIGGER)", sql_segment):
                        # 提取
                        create_pattern = r"CREATE\s+(?:/[*!].*?[*]/\s*)?(?:DEFINER\s*=\s*`[^`]+`@`[^`]+`\s*)?(TRIGGER|PROCEDURE|FUNCTION|VIEW|EVENT)\s+`([a-zA-Z0-9_]+)`"
                        match = re.search(create_pattern, sql_segment)
                        if match:
                            _type = match.group(1)  # 类型
                            _name = match.group(2)  # 名称
                            thread = threading.Thread(target=execute_query,
                                                      args=(server, section, sql_segment, _type, _name, db_name),
                                                      name=f"{section}-Thread")
                            thread.start()
                            threads.append((thread, server, section))

                    else:
                        for statement in sqlparse.parse(sql_segment):
                            # 创建线程执行每个 SQL 语句
                            sql_str = str(statement)  # 使用 sql_str 而不是 sql_query
                            # logging.info(f"执行SQL语句：{sql_str}")
                            print(f"执行到：{sql_str}")
                            thread = threading.Thread(target=execute_query,
                                                      args=(server, section, sql_str, 'Sql语句', None, None),
                                                      name=f"{section}-Thread")
                            thread.start()
                            threads.append((thread, server, section))
            for thread, server, section in threads:
                thread.join()
                logging.info(f"{thread.name} 已执行!")

            while not self.result_queue.empty():
                sql_query, result_df = self.result_queue.get()
                print(result_df)
                if not result_df.index.is_unique:
                    print("Warning: Duplicate index in result_df!")
                    result_df = result_df.reset_index(drop=True)
                if sql_query in result_dict:
                    result_dict[sql_query] = pd.concat([result_dict[sql_query], result_df], ignore_index=True)
                else:
                    result_dict[sql_query] = result_df.reset_index(drop=True)

            # 计算总耗时并显示
            elapsed_time = datetime.datetime.now() - start_time
            formatted_time = elapsed_time.total_seconds()
            self.timeLabel.setText(f"{round(formatted_time, 4)} Seconds")
            # 在 UI 上显示结果
            for sql_query, result_df in result_dict.items():
                print(1)
                self.result_ready.emit(result_df, sql_query)

        except Exception as e:
            logging.error(f"执行操作未知错误: {str(e)}")
            popup_manager.message_signal.emit(f"未完成执行的错误:{e}")


class Thread_2(QThread):
    signal = pyqtSignal()  # 括号里填写信号传递的参数
    stop_click = pyqtSignal(pd.DataFrame)

    def __init__(self, server_name, timeLabel, execute_group, execute_pms, execute_mem):
        super(Thread_2, self).__init__()
        self.timeLabel = timeLabel
        self.server_name = server_name
        self.result_queue = queue.Queue()
        self.execute_group = execute_group
        self.execute_pms = execute_pms
        self.execute_mem = execute_mem

    def run(self):
        try:
            server_name = self.server_name
            result_queue = queue.Queue()  # 创建一个队列来存储每个线程的执行结果
            config = parse_config(server_name)
            start_time = datetime.datetime.now()  # 记录开始时间

            def execute_query(server, section, result_queue):
                connection = connect_to_server(server)
                if connection:
                    try:
                        result = kill_sql(connection, section, None)
                        if result is not None:
                            result_queue.put(result)  # 将结果放入队列
                    except Exception as e:
                        logging.error(f"未完成执行的错误：=>>  {server}-{e}")
                        # popup_manager.message_signal.emit(f"未完成执行的错误：=>>  {server}-{e}")
                    finally:
                        connection.close()

            threads = []
            for section in config.sections():
                if section == 'group' and not self.execute_group.isChecked():
                    continue
                if section == 'member' and not self.execute_mem.isChecked():
                    continue
                server = config[section]
                thread = threading.Thread(target=execute_query, args=(server, section, result_queue),
                                          name=f"{section}-Thread")
                thread.start()
                threads.append((thread, server, section))
            for thread, server, section in threads:
                thread.join()
                logging.info(f"{thread.name} 已执行!")
            elapsed_time = datetime.datetime.now() - start_time
            formatted_time = elapsed_time.total_seconds()
            self.timeLabel.setText(f"{round(formatted_time, 4)} Seconds")

            # 在界面上显示结果
            result_df = pd.DataFrame()
            while not result_queue.empty():
                result = result_queue.get()
                if result_df.empty:
                    result_df = result
                else:
                    result_df = result_df.append(result, ignore_index=True)
            self.stop_click.emit(result_df)
        except Exception as e:
            logging.error(f"停止操作未知错误: {str(e)}")


class MainWindow(QtWidgets.QMainWindow, Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)

    def closeEvent(self, event):
        # 创建确认退出的对话框
        reply = QtWidgets.QMessageBox.question(
            self,
            '确认退出',
            '您确定要退出吗？',
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
            QtWidgets.QMessageBox.No
        )

        if reply == QtWidgets.QMessageBox.Yes:
            event.accept()
        else:
            event.ignore()

class LoadDatabasesThread(QThread):
    """后台线程：用于异步加载数据库"""
    databasesLoaded = pyqtSignal(str, list)  # 发射信号，返回实例名和数据库列表
    def __init__(self, instance_name ,get_server_config_fn):
        super().__init__()
        self.instance_name = instance_name
        self.get_server_config_fn = get_server_config_fn
    def run(self):
        """在后台线程中加载数据库列表"""
        print("加载")
        databases = self.get_databases_for_instance(self.instance_name)
        self.databasesLoaded.emit(self.instance_name, databases)  # 发射信号
    def get_databases_for_instance(self,  instance):
        """根据目标实例获取数据库列表"""
        server_config = self.get_server_config_fn(instance)
        if server_config:
            connection = connect_to_server(server_config)
            if connection:
                cursor = connection.cursor()
                cursor.execute("SHOW DATABASES;")
                databases = [db[0] for db in cursor.fetchall()]
                cursor.close()
                connection.close()
                return databases

class CopyTableto_Dialog(QDialog):
    """复制表到其他数据库/服务器"""
    def __init__(self, server_name, execute_mem, execute_group, parent=None):
        super(CopyTableto_Dialog, self).__init__(parent)
        self.execute_mem = execute_mem
        self.execute_group = execute_group
        self.server_name = server_name
        self.result_queue = queue.Queue()
        self.setWindowTitle("复制数据库")
        self.resize(1080, 900)
        layout = QHBoxLayout(self)
        self.setup_ui(layout)
    def setup_ui(self, layout):
        # 左侧: 源数据库选择
        left_layout = QVBoxLayout()
        self.source_db_label = QLabel("源实例:")
        self.source_db_combobox = QComboBox()
        self.source_table_list = QListWidget()
        self.source_table_list.setSelectionMode(QListWidget.MultiSelection)
        self.source_db_combobox.setCurrentIndex(0)
        self.source_db_combobox.currentTextChanged.connect(self.load_db)
        self.dblabel = QLabel("源数据库:")
        self.min_layout = QtWidgets.QHBoxLayout()

        self.db = QComboBox()
        self.load_table_btn = QPushButton("查询表")
        self.load_table_btn.clicked.connect(self.load_table_btn_a)

        self.min_layout.addWidget(self.db)
        self.min_layout.addWidget(self.load_table_btn)
        min_widget = QWidget()
        min_widget.setLayout(self.min_layout)
        # 添加搜索框
        self.search_box = QLineEdit()
        self.search_box.setPlaceholderText("输入表名搜索...")
        self.search_box.textChanged.connect(self.filter_tables)

        left_layout.addWidget(self.source_db_label)
        left_layout.addWidget(self.source_db_combobox)
        left_layout.addWidget(self.dblabel)
        left_layout.addWidget(min_widget)
        left_layout.addWidget(QLabel("选择要复制的表"))
        left_layout.addWidget(self.source_table_list)
        left_layout.addWidget(self.search_box)  # 添加搜索框
        # 右侧: 目标实例与数据库选择
        right_layout = QVBoxLayout()
        self.target_instance_label = QLabel("选择目标实例:")
        self.target_instance_list = QListWidget()
        self.target_instance_list.setSelectionMode(QListWidget.MultiSelection)
        self.target_db_container = QVBoxLayout()  # 用于容纳动态生成的目标数据库选择框
        self.copy_button = QPushButton("复制表数据")
        self.copy_button.clicked.connect(self.copy_data)

        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        self.progress_bar.setVisible(False)  # 初始时隐藏进度条

        right_layout.addWidget(self.target_instance_label)
        right_layout.addWidget(self.target_instance_list)
        right_layout.addLayout(self.target_db_container)
        right_layout.addWidget(self.copy_button)
        # 添加左右布局
        layout.addLayout(left_layout, 1)
        layout.addLayout(right_layout, 1)

        self.load_databases()
        self.target_instance_list.itemSelectionChanged.connect(self.load_target_databases)
    def load_databases(self):
        """解析配置文件并加载数据库名称到ComboBox"""
        config = parse_config(self.server_name)
        if config is None:
            self.show_error("无法解析配置文件！")
            return
        for section in config.sections():
            self.source_db_combobox.addItem(section)
            self.target_instance_list.addItems([section])
    def filter_tables(self):
        search_text = self.search_box.text().lower()  # 获取搜索框文本并转为小写
        for index in range(self.source_table_list.count()):
            item = self.source_table_list.item(index)
            item.setHidden(search_text not in item.text().lower())

    def load_db(self):
        server = self.source_db_combobox.currentText()
        server_config = self.get_server_config(server)
        if server_config:
            connection = connect_to_server(server_config)
            if connection:
                cursor = connection.cursor()
                cursor.execute("SHOW DATABASES;")
                tables = cursor.fetchall()
                self.db.clear()
                for table in tables:
                    table_name = table[0]  # 获取元组中的第一个元素，假设它是表名
                    self.db.addItem(table_name)
                cursor.close()
                connection.close()
        else:
            logging.error(f"未找到服务器配置：{server}")
            popup_manager.message_signal.emit(f"未找到服务器配置：{server}")

    def load_table_btn_a(self):
        self.load_tables()

    def load_tables(self):
        """加载源数据库中的表"""
        server = self.source_db_combobox.currentText()
        db = self.db.currentText()
        server_config = self.get_server_config(server)
        if server_config:
            connection = connect_to_server(server_config)
            if connection:
                cursor = connection.cursor()
                cursor.execute(f"SHOW TABLES FROM {db};")
                tables = cursor.fetchall()
                self.source_table_list.clear()
                for table in tables:
                    table_name = table[0]
                    self.source_table_list.addItem(table_name)
                cursor.close()
                connection.close()
        else:
            logging.error(f"未找到服务器配置：{server}")
            popup_manager.message_signal.emit(f"未找到服务器配置：{server}")


    def get_server_config(self, server_name):
        """根据实例名称获取对应的服务器配置"""
        config = parse_config(self.server_name)  # 这里是你解析配置文件的函数
        if config and server_name in config:
            return config[server_name]  # 返回对应的配置字典
        return None  # 如果没有找到该实例配置，返回None

    def get_selected_tables(self):
        selected_tables = []
        for index in range(self.source_table_list.count()):
            item = self.source_table_list.item(index)
            checkbox = self.source_table_list.itemWidget(item)  # 获取复选框
            if checkbox.isChecked():
                selected_tables.append(item.text())  # 如果复选框被选中，则记录表名
        return selected_tables
    def load_target_databases(self):
        """根据选中的目标实例加载目标数据库"""
        print("选中")
        selected_instances = [item.text() for item in self.target_instance_list.selectedItems()]
        # # 清空之前的数据库选择框
        for i in reversed(range(self.target_db_container.count())):
            widget = self.target_db_container.itemAt(i).widget()
            if widget:
                widget.deleteLater()
                # 显示进度条，准备开始加载
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        # 异步加载每个实例的数据库
        for instance in selected_instances:
            self.start_loading_databases(instance)

    def start_loading_databases(self, instance):
        """启动后台线程加载数据库"""
        thread = LoadDatabasesThread(instance, self.get_server_config)
        thread.databasesLoaded.connect(self.on_databases_loaded)
        thread.start()

    def on_databases_loaded(self, instance, databases):
        """当数据库加载完毕后更新UI"""
        self.progress_bar.setValue(self.progress_bar.value() + 20)  # 更新进度条
        label = QLabel(f"选择 {instance} 的数据库:")
        db_combobox = QComboBox()
        db_combobox.addItems(databases)
        self.target_db_container.addWidget(label)
        self.target_db_container.addWidget(db_combobox)

        # 当所有实例的数据库加载完毕时，隐藏进度条
        if self.progress_bar.value() >= 100:
            self.progress_bar.setVisible(False)


    def copy_data(self):
        """复制数据到目标数据库"""
        source_db = self.source_db_combobox.currentText()
        print(f"源实例：{source_db}")
        db = self.db.currentText()
        print(f"源数据库：{db}")
        server_config = self.get_server_config(source_db)
        selected_tables = [item.text() for item in self.source_table_list.selectedItems()]
        print(f"拷贝表：{selected_tables}")
        if not source_db or not selected_tables:
            self.show_error("请选择源数据库和至少一个表")
            return

        target_databases = []
        selected_instances = [item.text() for item in self.target_instance_list.selectedItems()]
        for i, instance in enumerate(selected_instances):
            j = 2 * (i + 1) - 1
            widget = self.target_db_container.itemAt(j).widget()
            if isinstance(widget, QComboBox):
                selected_db = widget.currentText()
                target_databases.append((instance, selected_db))
        if not target_databases:
            self.show_error("请选择目标数据库")
            return

        try:
            # 获取源数据库的数据
            connection = connect_to_server(server_config)
            cursor = connection.cursor()
            data = {}
            for table in selected_tables:
                cursor.execute(f"SELECT * FROM {db}.{table}")
                data[table] = cursor.fetchall()
                columns = [desc[0] for desc in cursor.description]
                data[table] = (columns, data[table])
            cursor.close()
            connection.close()

            # 使用线程池并行拷贝数据到目标数据库
            with ThreadPoolExecutor(max_workers=4) as executor:  # 设定适当的线程数
                futures = []
                for target_server_config, target_db in target_databases:
                    for table, (columns, rows) in data.items():
                        futures.append(
                            executor.submit(self.insert_data_to_target, target_server_config, target_db, table, columns,
                                            rows))
                # 等待所有任务完成
                for future in futures:
                    future.result()

        except Exception as e:
            print(f"数据复制失败: {e}")

    def insert_data_to_target(self, target_server_config, target_db, table_name, columns, data):
        """将数据插入到目标数据库"""
        print("复制")
        server_config = self.get_server_config(target_server_config)
        connection = connect_to_server(server_config)
        cursor = connection.cursor()
        try:
            # 检查表是否存在
            try:
                cursor.execute(f"SHOW TABLES LIKE '{table_name}'")
                if cursor.fetchone():
                    print(f"表 {table_name} 已存在")
                else:
                    print(f"表 {table_name} 不存在，创建表...")
                    create_table_query = self.get_create_table_query(target_server_config, target_db, table_name)
                    cursor.execute(create_table_query)
                    connection.commit()

            except pymysql.MySQLError as e:
                print(f"检查表存在失败: {e}")
                raise

            batch_size = 5000  # 批量插入大小增加
            placeholders = ", ".join(["%s"] * len(columns))
            insert_query = f"INSERT INTO {target_db}.{table_name} ({', '.join(columns)}) VALUES ({placeholders})"

            print("开始插入数据...")
            for i in range(0, len(data), batch_size):
                batch = data[i:i + batch_size]
                cursor.executemany(insert_query, batch)
                if i % (batch_size * 500) == 0:  # 每10个批次提交一次
                    connection.commit()
                    print(f"插入了 {min(i + batch_size, len(data))} 条数据...")

            connection.commit()  # 最后一次提交
            print(f"数据插入成功到 {target_db}.{table_name}")

        except Exception as e:
            connection.rollback()
            print(f"插入数据失败到 {target_db}: {e}")
        finally:
            cursor.close()
            connection.close()

    def get_create_table_query(self, target_server_config, target_db, table_name):
        """从源数据库获取创建表的 SQL 语句"""
        print(f"{target_server_config}")
        server_config = self.get_server_config(target_server_config)
        db = self.db.currentText()
        try:
            source_connection = connect_to_server(server_config)
            source_cursor = source_connection.cursor()

            # 尝试获取源数据库的表创建语句
            try:
                source_cursor.execute(f"SHOW CREATE TABLE {db}.{table_name}")
                create_table_result = source_cursor.fetchone()
                if create_table_result:
                    create_table_query = create_table_result[1]
                    create_table_query = create_table_query.replace(f"CREATE TABLE `{table_name}`",
                                                                    f"CREATE TABLE `{target_db}`.`{table_name}`")
                    print(f"创建表 SQL: {create_table_query}")
                else:
                    raise Exception(f"无法获取表 {table_name} 的创建 SQL")
            except pymysql.MySQLError as e:
                if e.args[0] == 1146:  # 1146错误代码表明表不存在
                    raise Exception(f"表 {table_name} 在源数据库中不存在，无法获取创建 SQL")
                else:
                    raise  # 其他 MySQL 错误继续抛出

            source_cursor.close()
            source_connection.close()

            return create_table_query

        except Exception as e:
            self.show_error(f"获取创建表 SQL 失败: {e}")
            raise

    def show_error(self, message):
        """显示错误消息"""
        QMessageBox.critical(self, "错误", message)
    def show_info(self, message):
        """显示信息消息"""
        QMessageBox.information(self, "信息", message)

class ProcessDialog(QDialog):
    """mysql进程列表"""

    def __init__(self, server_name, execute_mem, execute_group, parent=None):
        super(ProcessDialog, self).__init__(parent)
        self.execute_mem = execute_mem
        self.execute_group = execute_group
        self.server_name = server_name
        self.result_queue = queue.Queue()
        self.setWindowTitle("进程列表")
        self.resize(1080, 600)
        layout = QVBoxLayout(self)

        self.label = QLabel("进程：")
        layout.addWidget(self.label)
        self.process_table = QTableWidget()
        layout.addWidget(self.process_table)

        self.refresh_button = QPushButton("刷新")
        self.refresh_button.clicked.connect(self.load_processes)
        layout.addWidget(self.refresh_button)

        self.kill_button = QPushButton("杀了它")
        self.kill_button.clicked.connect(self.kill_selected_process)
        layout.addWidget(self.kill_button)

        self.load_processes()

        self.timer = QTimer(self)
        self.timer.timeout.connect(self.load_processes)
        self.timer.start(10000)  # 10秒刷新

    def load_processes(self):
        self.process_table.clear()
        self.result_queue = queue.Queue()
        try:
            config = parse_config(self.server_name)

            def execute_query(server, section, sql):
                connection = connect_to_server(server)
                cursor = connection.cursor()
                if connection:
                    try:
                        logging.info(f"执行SQL语句：{sql}")
                        cursor.execute(sql)
                        result = cursor.fetchall()
                        column_names = [i[0] for i in cursor.description]
                        if result is not None:
                            self.result_queue.put((section, result, column_names))  # 包含section和列名
                    except Exception as e:
                        logging.error(f"未完成执行的错误：=>>  {server}-{e}")
                    finally:
                        cursor.close()
                        connection.close()

            threads = []
            for section in config.sections():
                if section == 'group' and not self.execute_group.isChecked():
                    continue
                if section == 'member' and not self.execute_mem.isChecked():
                    continue
                server = config[section]
                sql_str = "SHOW PROCESSLIST"
                thread = threading.Thread(target=execute_query, args=(server, section, sql_str),
                                          name=f"{section}-Thread")
                thread.start()
                threads.append((thread, section))

            for thread, section in threads:
                thread.join()
                logging.info(f"{thread.name} 已执行!")

            self.process_table.setRowCount(0)
            headers_set = False
            while not self.result_queue.empty():
                section, result, column_names = self.result_queue.get()
                if not headers_set:
                    self.process_table.setColumnCount(len(column_names) + 1)
                    self.process_table.setHorizontalHeaderLabels(["服务器节"] + column_names)
                    headers_set = True

                for row in result:
                    row_count = self.process_table.rowCount()
                    self.process_table.insertRow(row_count)
                    self.process_table.setItem(row_count, 0, QTableWidgetItem(section))
                    for column_index, value in enumerate(row):
                        self.process_table.setItem(row_count, column_index + 1, QTableWidgetItem(str(value)))

        except Exception as e:
            logging.error(f"执行操作未知错误: {str(e)}")

    def kill_selected_process(self):
        selected_items = self.process_table.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "警告", "搞毛，没选到！")
            return

        process_id = None
        server_section = None
        selected_row = selected_items[0].row()

        for col in range(self.process_table.columnCount()):
            item = self.process_table.item(selected_row, col)
            if item:
                if col == 0:  # 服务器节
                    server_section = item.text()
                elif col == 1:  # 进程ID
                    process_id = item.text()

        if process_id is None or server_section is None:
            QMessageBox.warning(self, "警告", "搞毛，没选到！")
            return

        kill_sql = f"KILL {process_id};"
        config = parse_config(self.server_name)
        # 执行杀掉进程的操作
        for section in config.sections():
            if section == server_section:
                server = config[section]
                connection = connect_to_server(server)
                cursor = connection.cursor()
                try:
                    logging.info(f"执行 KILL SQL：{kill_sql}")
                    cursor.execute(kill_sql)
                    connection.commit()
                    QMessageBox.information(self, "成功", f"成功杀掉进程 {process_id} 在服务器节 {server_section}")
                    self.load_processes()
                except Exception as e:
                    logging.error(f"杀掉进程时发生错误：{e}")
                    QMessageBox.critical(self, "错误", f"无法杀掉进程 {process_id}：{e}")
                finally:
                    cursor.close()
                    connection.close()


class ServerDialog(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("添加服务器")

        # 创建表单布局
        layout = QFormLayout()

        # 添加输入字段
        self.service_group_name = QLineEdit()
        self.host = QLineEdit()
        self.user = QLineEdit()
        self.password = QLineEdit()
        self.port = QLineEdit()
        self.database = QLineEdit()
        self.ssh_checkbox = QCheckBox("SSH")

        # SSH相关字段
        self.ssh_host = QLineEdit()
        self.ssh_user = QLineEdit()
        self.ssh_password = QLineEdit()  # 确保这是一个新的QLineEdit实例
        self.ssh_port = QLineEdit()

        # 密码可见性按钮
        self.password_visible_button = QPushButton("显示")
        self.password_visible_button.setCheckable(True)
        self.password_visible_button.toggled.connect(self.toggle_password_visibility)

        self.ssh_password_visible_button = QPushButton("显示")
        self.ssh_password_visible_button.setCheckable(True)
        self.ssh_password_visible_button.toggled.connect(self.toggle_ssh_password_visibility)

        # 把SSH字段隐藏，初始状态
        self.toggle_ssh_fields(False)  # 初始状态为禁用

        # 连接复选框状态变化信号
        self.ssh_checkbox.toggled.connect(self.toggle_ssh_fields)

        # 添加到布局
        layout.addRow(QLabel("服务组名称:"), self.service_group_name)
        layout.addRow(QLabel("主机:"), self.host)
        layout.addRow(QLabel("用户:"), self.user)

        layout.addRow(QLabel("密码:"), self.create_password_layout())  # 使用创建密码布局的函数
        layout.addRow(QLabel("端口:"), self.port)
        layout.addRow(QLabel("数据库:"), self.database)
        layout.addRow(self.ssh_checkbox)
        layout.addRow(QLabel("SSH主机:"), self.ssh_host)
        layout.addRow(QLabel("SSH用户:"), self.ssh_user)
        layout.addRow(QLabel("SSH密码:"), self.create_ssh_password_layout())  # 使用创建SSH密码布局的函数
        layout.addRow(QLabel("SSH端口:"), self.ssh_port)

        # 添加确认按钮
        self.ok_button = QPushButton("确定")
        self.ok_button.clicked.connect(self.handle_ok_click)
        layout.addWidget(self.ok_button)

        self.setLayout(layout)

    def create_password_layout(self):
        """创建密码输入框和按钮的布局"""
        layout = QHBoxLayout()
        self.password.setEchoMode(QLineEdit.Password)
        layout.addWidget(self.password)
        layout.addWidget(self.password_visible_button)
        return layout

    def create_ssh_password_layout(self):
        """创建SSH密码输入框和按钮的布局"""
        layout = QHBoxLayout()
        self.password.setEchoMode(QLineEdit.Password)
        layout.addWidget(self.ssh_password)
        layout.addWidget(self.ssh_password_visible_button)
        return layout

    def toggle_ssh_fields(self, checked):
        """Toggle the SSH input fields based on checkbox state."""
        state = not checked
        self.ssh_host.setDisabled(state)
        self.ssh_user.setDisabled(state)
        self.ssh_password.setDisabled(state)
        self.ssh_port.setDisabled(state)

    def toggle_password_visibility(self, checked):
        """切换主密码可见性"""
        if checked:
            self.password.setEchoMode(QLineEdit.Normal)
            self.password_visible_button.setText("隐藏")
        else:
            self.password.setEchoMode(QLineEdit.Password)
            self.password_visible_button.setText("显示")

    def toggle_ssh_password_visibility(self, checked):
        """切换SSH密码可见性"""
        if checked:
            self.ssh_password.setEchoMode(QLineEdit.Normal)
            self.ssh_password_visible_button.setText("隐藏")
        else:
            self.ssh_password.setEchoMode(QLineEdit.Password)
            self.ssh_password_visible_button.setText("显示")

    def handle_ok_click(self):
        """处理确定按钮点击事件"""
        if self.validate_inputs():
            self.accept()  # 输入有效，接受对话框

    def validate_inputs(self):
        """验证主机和端口的输入"""
        host_text = self.host.text().strip()
        port_text = self.port.text().strip()

        # 验证主机地址（IP 格式或域名）
        ip_pattern = r'^(25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.(25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.(25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.(25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)$'
        domain_pattern = r'^[a-zA-Z0-9][-a-zA-Z0-9.]*[a-zA-Z0-9]$'  # 更严格的域名正则

        if not re.match(ip_pattern, host_text) and not re.match(domain_pattern, host_text):
            QMessageBox.warning(self, "输入错误", "请输入有效的主机地址（IP 或 域名）。")
            return False

        # 验证端口号
        if not port_text.isdigit():
            QMessageBox.warning(self, "输入错误", "端口号必须是数字。")
            return False

        port_number = int(port_text)
        if not (1 <= port_number <= 65535):
            QMessageBox.warning(self, "输入错误", "请输入有效的端口号（1-65535）。")
            return False

        return True


class LoginWindow(QtWidgets.QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle('登录界面')

        # 设置窗口大小和初始位置
        self.setGeometry(500, 300, 400, 300)
        self.setStyleSheet("background-color: #f5f5f5;")  # 设置背景色

        # 使用 QVBoxLayout 来更好的控制布局
        layout = QtWidgets.QVBoxLayout(self)

        # 创建一个标题标签
        self.title_label = QtWidgets.QLabel('欢迎登录', self)
        self.title_label.setAlignment(QtCore.Qt.AlignCenter)
        self.title_label.setFont(QtGui.QFont('Arial', 20, QtGui.QFont.Bold))
        layout.addWidget(self.title_label)

        layout.addSpacing(30)  # 加一些间距，使布局更舒服

        # 用户名
        self.label_username = QtWidgets.QLabel('用户名:', self)
        self.label_username.setFont(QtGui.QFont('Arial', 12))
        layout.addWidget(self.label_username)

        self.username_input = QtWidgets.QLineEdit(self)
        self.username_input.setPlaceholderText('请输入用户名')
        self.username_input.setStyleSheet("""
            QLineEdit {
                background-color: white;
                border: 1px solid #ccc;
                border-radius: 10px;
                padding: 10px;
                font-size: 14px;
            }
        """)
        layout.addWidget(self.username_input)

        layout.addSpacing(20)  # 再加些间距

        # 密码
        self.label_password = QtWidgets.QLabel('密码:', self)
        self.label_password.setFont(QtGui.QFont('Arial', 12))
        layout.addWidget(self.label_password)

        self.password_input = QtWidgets.QLineEdit(self)
        self.password_input.setPlaceholderText('请输入密码')
        self.password_input.setEchoMode(QtWidgets.QLineEdit.Password)  # 密码框隐藏输入内容
        self.password_input.setStyleSheet("""
            QLineEdit {
                background-color: white;
                border: 1px solid #ccc;
                border-radius: 10px;
                padding: 10px;
                font-size: 14px;
            }
        """)
        layout.addWidget(self.password_input)

        layout.addSpacing(30)  # 再加些间距

        # 登录按钮
        self.login_button = QtWidgets.QPushButton('登录', self)
        self.login_button.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                padding: 12px 20px;
                text-align: center;
                font-size: 16px;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        layout.addWidget(self.login_button)

        layout.addStretch(1)  # 让布局保持一定的对齐

        # 按钮点击事件
        self.login_button.clicked.connect(self.handle_login)

        # 支持通过回车键登录
        self.username_input.returnPressed.connect(self.handle_login)
        self.password_input.returnPressed.connect(self.handle_login)

    def handle_login(self):
        username = self.username_input.text()
        password = self.password_input.text()

        if username == 'admin' and password == 'password':
            self.accept_login()
        else:
            QtWidgets.QMessageBox.warning(self, '登录失败', '用户名或密码错误')

    def accept_login(self):
        # 登录成功，关闭当前窗口，打开主窗口
        self.close()
        self.main_window = MainWindow()
        self.main_window.show()


if __name__ == "__main__":
    import sys
    from PyQt5 import QtWidgets
    app = QtWidgets.QApplication(sys.argv)
    login_window = LoginWindow()
    login_window.show()
    sys.exit(app.exec_())
